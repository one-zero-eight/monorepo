"""Build styled XLSX schedule exports for one section (groups weekly / calendar)."""

import datetime as dtm
import io
import re
from collections import Counter, defaultdict
from colorsys import hls_to_rgb
from dataclasses import dataclass
from typing import Literal

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schedule_assistant.modules.distributions.mapping import iter_section_group_codes
from src.schedule_assistant.modules.issues.booking_slots import (
    _edit_for_meeting_date,
    _weekly_meeting_dates_in_window,
)
from src.schedule_assistant.modules.schedule_config.schemas import (
    CourseConfig,
    ScheduleConfig,
    SectionConfig,
    SectionsConfig,
    TermConfig,
    TermTimeSlot,
)
from src.schedule_assistant.modules.schedule_config.semester_windows import (
    resolve_audience_semester,
    union_semester_window,
)
from src.schedule_assistant.modules.schedule_config.validation import build_selector_map, expand_group_tokens
from src.schedule_assistant.weekday import Weekday, week_start_for_date

ExportLayout = Literal["groups", "compact_groups", "calendar"]

HEADER_FILL = PatternFill("solid", fgColor="548DD4")
TIME_HEADER_FILL = PatternFill("solid", fgColor="741B47")
WEEKDAY_FILL = PatternFill("solid", fgColor="00B0F0")
WEEK_HEADER_FILL = PatternFill("solid", fgColor="FCE5CD")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="000000")
THICK = Side(style="thick", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="000000")
WEEKDAY_FONT = Font(bold=True, color="000000")
TITLE_FONT = Font(bold=True)
NORMAL_FONT = Font(bold=False)
COMPACT_TITLE_INLINE_FONT = InlineFont(b=True, sz=11)
COMPACT_DETAIL_INLINE_FONT = InlineFont(b=False, sz=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

GROUPS_TIME_COL_WIDTH = 14
GROUPS_COL_WIDTH = 24
GROUPS_EVENT_ROW_HEIGHT_MIN = 15
GROUPS_LINE_HEIGHT = 15
COMPACT_GROUPS_TIME_COL_WIDTH = 12
COMPACT_GROUPS_COL_WIDTH = 13
COMPACT_GROUPS_ROW_HEIGHT_MIN = 30
COMPACT_GROUPS_LINE_HEIGHT = 12
COMPACT_WEEKDAY_LABELS = {
    Weekday.MONDAY: "Mon",
    Weekday.TUESDAY: "Tue",
    Weekday.WEDNESDAY: "Wed",
    Weekday.THURSDAY: "Thu",
    Weekday.FRIDAY: "Fri",
    Weekday.SATURDAY: "Sat",
    Weekday.SUNDAY: "Sun",
}
DISTRIBUTIONS_SHEET_NAME = "Distributions"
DISTRIBUTIONS_HEADERS = ("E-mail", "Group", "Section")
DISTRIBUTIONS_COL_WIDTHS = (36, 28, 18)
INSTRUCTORS_SHEET_NAME = "Instructors"
INSTRUCTORS_HEADERS = ("Name EN", "Name RU", "Email", "Alias", "Position", "Courses", "ID")
INSTRUCTORS_COL_WIDTHS = (28, 28, 32, 16, 22, 48, 18)
SUBJECTS_SHEET_NAME = "Subjects"
SUBJECTS_HEADERS = ("Section", "Short name", "Name", "Name RU", "Groups", "Instructors")
SUBJECTS_COL_WIDTHS = (18, 14, 36, 36, 42, 48)
CALENDAR_LEGEND_HEADERS = ("Short name", "Course Name", "Instructor")
CALENDAR_LEGEND_COL_WIDTHS = (14, 42, 28)

_INVALID_SHEET_CHARS = re.compile(r"[\[\]\*\/\\?:]")


@dataclass(frozen=True)
class ExportMeeting:
    course: str
    course_short_name: str | None
    tag: str
    groups: tuple[str, ...]
    date: dtm.date
    start: dtm.time
    end: dtm.time
    room: str
    instructors: tuple[str, ...]


@dataclass(frozen=True)
class ExportColumn:
    year_label: str
    group_id: str
    group_label: str


@dataclass(frozen=True)
class ProgramBlock:
    year_label: str
    columns: list[ExportColumn]
    time_col: int
    group_start: int
    group_end: int


def sanitize_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", name).strip() or "Schedule"
    return cleaned[:31]


_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitize_filename(name: str) -> str:
    cleaned = _INVALID_FILENAME_CHARS.sub(" ", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .") or "Schedule"
    return cleaned[:180]


def export_filename(term_name: str) -> str:
    term = sanitize_filename(term_name)
    return f"{term or 'Schedule'}.xlsx"


def section_export_layout(section: SectionConfig) -> ExportLayout:
    if section.default_layout == "calendar":
        return "calendar"
    if section.default_layout == "compact_groups":
        return "compact_groups"
    return "groups"


def _unique_sheet_name(base: str, used: set[str]) -> str:
    name = sanitize_sheet_name(base)
    if name not in used:
        return name
    for index in range(2, 1000):
        suffix = f" {index}"
        candidate = sanitize_sheet_name(f"{base[: max(1, 31 - len(suffix))]}{suffix}")
        if candidate not in used:
            return candidate
    return sanitize_sheet_name(f"{base[:27]}…")


def _hhmm(value: dtm.time) -> str:
    return value.strftime("%H:%M")


def _slot_label(start: dtm.time, end: dtm.time) -> str:
    return f"{_hhmm(start)}-{_hhmm(end)}"


def _instructor_ids(value: str | list[str] | None) -> tuple[str, ...]:
    if value is None:
        return ()
    if isinstance(value, list):
        return tuple(item for item in value if item)
    return (value,) if value else ()


def _normalize_tracks(program: SectionConfig.SectionProgram) -> list[tuple[str, list[str]]]:
    if program.tracks:
        return [(track.name, list(track.groups)) for track in program.tracks if track.groups]
    if program.groups:
        return [(program.code or program.name or "—", list(program.groups))]
    return []


def section_group_set(section: SectionConfig) -> set[str]:
    groups: set[str] = set()
    for program in section.programs:
        groups.update(program_group_set(program))
    return groups


def program_group_set(program: SectionConfig.SectionProgram) -> set[str]:
    groups: set[str] = set()
    for _track_name, track_groups in _normalize_tracks(program):
        groups.update(track_groups)
    return groups


def filter_meetings_for_groups(
    meetings: list[ExportMeeting],
    groups: set[str],
) -> list[ExportMeeting]:
    if not groups:
        return []
    return [meeting for meeting in meetings if any(group in groups for group in meeting.groups)]


def _instructor_label_by_id(config: ScheduleConfig) -> dict[str, str]:
    out: dict[str, str] = {}
    for instructor in config.instructors:
        label = (instructor.name_en or instructor.name_ru or instructor.id or "").strip()
        out[instructor.id] = label or instructor.id
        if instructor.email:
            out[instructor.email] = label or instructor.email
    return out


def format_instructors(ids: tuple[str, ...], labels: dict[str, str]) -> str:
    return " / ".join(labels.get(item, item) for item in ids if item)


def meeting_title(meeting: ExportMeeting) -> str:
    course = meeting.course.strip() or "—"
    tag = meeting.tag.strip()
    if tag:
        return f"{course} ({tag})"
    return course


def pattern_signature(meeting: ExportMeeting) -> str:
    groups = "|".join(sorted(meeting.groups))
    instructors = "|".join(meeting.instructors)
    return f"{meeting.course}|{meeting.tag}|{_hhmm(meeting.start)}|{groups}|{instructors}|{meeting.room}"


def cell_signature(meetings: list[ExportMeeting]) -> str:
    if not meetings:
        return ""
    return "||".join(f"{pattern_signature(m)}#{i}" for i, m in enumerate(meetings))


def course_fill_color(course: str) -> str:
    key = course.strip() or "—"
    h = 0
    for ch in key:
        h = (h * 31 + ord(ch)) & 0xFFFFFFFF
    hue = ((h * 137.508) % 360.0) / 360.0
    mix = (h >> 3) & 0xFF
    sat = 0.45 + (mix % 4) * 0.08
    light = 0.78 + ((mix >> 2) % 5) * 0.03
    r, g, b = hls_to_rgb(hue, light, sat)
    return f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"


def expand_meetings(config: ScheduleConfig) -> list[ExportMeeting]:
    term = config.term
    selector_map = build_selector_map(SectionsConfig(sections=term.sections, students_groups=config.students_groups))
    meetings: list[ExportMeeting] = []

    for course in config.courses:
        short_name = (course.short_name or "").strip() or None
        for component in course.components:
            for session in component.sessions or []:
                tokens = session.audience or component.audience
                groups = tuple(sorted(expand_group_tokens(tokens, selector_map)))

                for occurrence in session.dates_pattern or []:
                    meetings.append(
                        ExportMeeting(
                            course=course.name,
                            course_short_name=short_name,
                            tag=str(component.tag),
                            groups=groups,
                            date=occurrence.date,
                            start=occurrence.start_time,
                            end=occurrence.end_time,
                            room=(occurrence.room or "").strip(),
                            instructors=_instructor_ids(occurrence.instructor),
                        )
                    )

                for slot in session.weekly_pattern or []:
                    if term.days and slot.weekday not in term.days:
                        continue
                    window = resolve_audience_semester(term, list(tokens))
                    if window is None:
                        continue
                    for meeting_date in _weekly_meeting_dates_in_window(window, slot.weekday):
                        edit = _edit_for_meeting_date(meeting_date, slot.edits or [], term)
                        if edit is not None and edit.cancel:
                            continue
                        resolved_date = edit.date if edit is not None and edit.date is not None else meeting_date
                        resolved_start = (
                            edit.start_time if edit is not None and edit.start_time is not None else slot.start_time
                        )
                        resolved_end = (
                            edit.end_time if edit is not None and edit.end_time is not None else slot.end_time
                        )
                        resolved_room = edit.room if edit is not None and edit.room is not None else slot.room
                        resolved_instructor = (
                            edit.instructor if edit is not None and edit.instructor is not None else slot.instructor
                        )
                        meetings.append(
                            ExportMeeting(
                                course=course.name,
                                course_short_name=short_name,
                                tag=str(component.tag),
                                groups=groups,
                                date=resolved_date,
                                start=resolved_start,
                                end=resolved_end,
                                room=(resolved_room or "").strip(),
                                instructors=_instructor_ids(resolved_instructor),
                            )
                        )
    return meetings


def filter_meetings_for_section(
    meetings: list[ExportMeeting],
    section: SectionConfig,
) -> list[ExportMeeting]:
    return filter_meetings_for_groups(meetings, section_group_set(section))


def build_columns(
    section: SectionConfig,
    meetings: list[ExportMeeting],
    config: ScheduleConfig,
    *,
    include_unused: bool = False,
) -> list[ExportColumn]:
    used_groups = {group for meeting in meetings for group in meeting.groups}
    group_names = {group.code: (group.name or group.code) for group in config.students_groups}
    columns: list[ExportColumn] = []
    seen: set[str] = set()

    for program in section.programs:
        year_label = program.name or section.code
        for _track_name, track_groups in _normalize_tracks(program):
            for group_id in track_groups:
                if group_id in seen:
                    continue
                if not include_unused and used_groups and group_id not in used_groups:
                    continue
                columns.append(
                    ExportColumn(
                        year_label=year_label,
                        group_id=group_id,
                        group_label=group_names.get(group_id, group_id),
                    )
                )
                seen.add(group_id)

    if not columns:
        for program in section.programs:
            year_label = program.name or section.code
            for _track_name, track_groups in _normalize_tracks(program):
                for group_id in track_groups:
                    if group_id in seen:
                        continue
                    columns.append(
                        ExportColumn(
                            year_label=year_label,
                            group_id=group_id,
                            group_label=group_names.get(group_id, group_id),
                        )
                    )
                    seen.add(group_id)
    return columns


def _compact_occupied_slots(
    meetings: list[ExportMeeting],
    term: TermConfig,
) -> list[tuple[Weekday, TermTimeSlot]]:
    slots = list(term.time_slots)
    slot_starts = {slot.start_time for slot in slots}
    occupied: set[tuple[Weekday, dtm.time]] = set()
    extra_slots: dict[dtm.time, TermTimeSlot] = {}
    for meeting in meetings:
        row_start = meeting.start if meeting.start in slot_starts else _nearest_slot_start(meeting.start, slots)
        if row_start is None:
            row_start = meeting.start
            extra_slots.setdefault(
                row_start,
                TermTimeSlot(start_time=meeting.start, end_time=meeting.end),
            )
        occupied.add((list(Weekday)[meeting.date.weekday()], row_start))

    ordered_slots = sorted(
        [*slots, *extra_slots.values()],
        key=lambda slot: (slot.start_time, slot.end_time),
    )
    return [(day, slot) for day in _term_days(term) for slot in ordered_slots if (day, slot.start_time) in occupied]


def _nearest_slot_start(start: dtm.time, slots: list[TermTimeSlot]) -> dtm.time | None:
    if not slots:
        return None
    best = min(
        slots,
        key=lambda slot: abs((slot.start_time.hour * 60 + slot.start_time.minute) - (start.hour * 60 + start.minute)),
    )
    return best.start_time


def _group_sizes(config: ScheduleConfig) -> dict[str, int | None]:
    out: dict[str, int | None] = {}
    for group in config.students_groups:
        if group.estimated_size is not None:
            out[group.code] = group.estimated_size
        elif group.students:
            out[group.code] = len(group.students)
        else:
            out[group.code] = None
    return out


def _collapse_weekly_cells(
    meetings: list[ExportMeeting],
    slots: list[TermTimeSlot],
) -> dict[tuple[Weekday, dtm.time, str], list[ExportMeeting]]:
    """Map (weekday, slot_start, group) -> representative meetings (most frequent patterns)."""
    buckets: dict[tuple[Weekday, dtm.time, str], list[ExportMeeting]] = defaultdict(list)
    slot_starts = {slot.start_time for slot in slots}

    for meeting in meetings:
        weekday = list(Weekday)[meeting.date.weekday()]
        row_start = meeting.start if meeting.start in slot_starts else _nearest_slot_start(meeting.start, slots)
        if row_start is None:
            continue
        for group_id in meeting.groups:
            buckets[(weekday, row_start, group_id)].append(meeting)

    result: dict[tuple[Weekday, dtm.time, str], list[ExportMeeting]] = {}
    for key, items in buckets.items():
        counts = Counter(pattern_signature(item) for item in items)
        by_sign: dict[str, ExportMeeting] = {}
        for item in items:
            sign = pattern_signature(item)
            by_sign.setdefault(sign, item)
        # Keep unique patterns ordered by frequency (desc), then signature
        ordered_signs = sorted(counts.keys(), key=lambda sign: (-counts[sign], sign))
        result[key] = [by_sign[sign] for sign in ordered_signs]
    return result


def _apply_fill(
    cell,
    fill: PatternFill,
    font: Font | None = None,
    align: Alignment | None = CENTER,
    border: Border | None = THIN_BORDER,
) -> None:
    cell.fill = fill
    if border is not None:
        cell.border = border
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align


def _triple_block_border(*, row_offset: int, col: int, start_col: int, end_col: int) -> Border:
    """Outer border around a 3-row event/time block; no internal grid lines."""
    return Border(
        left=THIN if col == start_col else None,
        right=THIN if col == end_col else None,
        top=THIN if row_offset == 0 else None,
        bottom=THIN if row_offset == 2 else None,
    )


def _section_block_border(*, row: int, col: int, start_row: int, end_row: int, start_col: int, end_col: int) -> Border:
    """Thick outer border for a section block; thin internal cell grid."""
    return Border(
        left=THICK if col == start_col else THIN,
        right=THICK if col == end_col else THIN,
        top=THICK if row == start_row else THIN,
        bottom=THICK if row == end_row else THIN,
    )


def _style_triple_block(
    ws: Worksheet,
    *,
    start_row: int,
    start_col: int,
    end_col: int,
    fill: PatternFill,
    fonts: tuple[Font | None, Font | None, Font | None],
) -> None:
    for row_offset in range(3):
        font = fonts[row_offset]
        r = start_row + row_offset
        for c in range(start_col, end_col + 1):
            _apply_fill(
                ws.cell(r, c),
                fill,
                font,
                CENTER,
                border=_triple_block_border(
                    row_offset=row_offset,
                    col=c,
                    start_col=start_col,
                    end_col=end_col,
                ),
            )


def _term_days(term: TermConfig) -> list[Weekday]:
    return list(term.days) if term.days else list(Weekday)[:6]


def _program_blocks(columns: list[ExportColumn]) -> list[ProgramBlock]:
    """One time column + group columns per program (year_label), like Core sheets."""
    blocks: list[ProgramBlock] = []
    i = 0
    excel_col = 1
    while i < len(columns):
        year = columns[i].year_label
        j = i + 1
        while j < len(columns) and columns[j].year_label == year:
            j += 1
        groups = columns[i:j]
        time_col = excel_col
        group_start = excel_col + 1
        group_end = excel_col + len(groups)
        blocks.append(
            ProgramBlock(
                year_label=year,
                columns=groups,
                time_col=time_col,
                group_start=group_start,
                group_end=group_end,
            )
        )
        excel_col = group_end + 1
        i = j
    return blocks


def _wrapped_line_count(text: str, width_chars: float) -> int:
    """Estimate wrapped line count for Excel wrap_text (word-aware)."""
    cleaned = (text or "").replace("\r\n", "\n").strip()
    if not cleaned:
        return 1
    chars_per_line = max(8, int(width_chars * 0.95))
    total = 0
    for paragraph in cleaned.split("\n"):
        words = paragraph.split()
        if not words:
            total += 1
            continue
        lines = 1
        line_len = 0
        for word in words:
            piece = len(word) if line_len == 0 else len(word) + 1
            if line_len + piece <= chars_per_line:
                line_len += piece
                continue
            lines += 1
            line_len = len(word)
            while line_len > chars_per_line:
                lines += 1
                line_len -= chars_per_line
        total += lines
    return max(1, total)


def _row_height_for_text(text: str, col_span: int) -> float:
    width = GROUPS_COL_WIDTH * max(1, col_span)
    return max(GROUPS_EVENT_ROW_HEIGHT_MIN, _wrapped_line_count(text, width) * GROUPS_LINE_HEIGHT)


def _write_event_block(
    ws: Worksheet,
    *,
    start_row: int,
    start_col: int,
    end_col: int,
    meetings: list[ExportMeeting],
    instructor_labels: dict[str, str],
) -> tuple[str, str, str]:
    """Write a 3-row event (or empty) block. Returns (title, instructors, room) texts."""
    if not meetings:
        for c in range(start_col, end_col + 1):
            _style_triple_block(
                ws,
                start_row=start_row,
                start_col=c,
                end_col=c,
                fill=WHITE_FILL,
                fonts=(None, None, None),
            )
        return ("", "", "")

    meeting = meetings[0]
    title = meeting_title(meeting)
    instructors = format_instructors(meeting.instructors, instructor_labels)
    room = meeting.room
    fill = PatternFill("solid", fgColor=course_fill_color(meeting.course))
    ws.cell(start_row, start_col, title)
    ws.cell(start_row + 1, start_col, instructors)
    ws.cell(start_row + 2, start_col, room)
    for r in range(start_row, start_row + 3):
        if end_col > start_col:
            ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
    _style_triple_block(
        ws,
        start_row=start_row,
        start_col=start_col,
        end_col=end_col,
        fill=fill,
        fonts=(TITLE_FONT, NORMAL_FONT, NORMAL_FONT),
    )
    return (title, instructors, room)


def _write_groups_sheet(
    ws: Worksheet,
    *,
    columns: list[ExportColumn],
    term: TermConfig,
    cells: dict[tuple[Weekday, dtm.time, str], list[ExportMeeting]],
    instructor_labels: dict[str, str],
    group_sizes: dict[str, int | None],
) -> None:
    days = _term_days(term)
    slots = list(term.time_slots)
    blocks = _program_blocks(columns)

    for block in blocks:
        # Time column header (merged over rows 1–2), like Core A1:A2
        time_header = ws.cell(1, block.time_col)
        _apply_fill(time_header, TIME_HEADER_FILL, HEADER_FONT, CENTER)
        ws.merge_cells(start_row=1, start_column=block.time_col, end_row=2, end_column=block.time_col)
        _apply_fill(ws.cell(2, block.time_col), TIME_HEADER_FILL, HEADER_FONT, CENTER)
        ws.column_dimensions[get_column_letter(block.time_col)].width = GROUPS_TIME_COL_WIDTH

        # Program / year label across group columns only
        year_cell = ws.cell(1, block.group_start, block.year_label)
        _apply_fill(year_cell, HEADER_FILL, HEADER_FONT, CENTER)
        if block.group_end > block.group_start:
            ws.merge_cells(
                start_row=1,
                start_column=block.group_start,
                end_row=1,
                end_column=block.group_end,
            )
            for c in range(block.group_start, block.group_end + 1):
                _apply_fill(ws.cell(1, c), HEADER_FILL, HEADER_FONT, CENTER)

        for offset, column in enumerate(block.columns):
            size = group_sizes.get(column.group_id)
            label = column.group_label
            if size is not None:
                label = f"{label} ({size})"
            cell = ws.cell(2, block.group_start + offset, label)
            _apply_fill(cell, HEADER_FILL, HEADER_FONT, CENTER)
            ws.column_dimensions[get_column_letter(block.group_start + offset)].width = GROUPS_COL_WIDTH

    row = 3
    for day in days:
        for block in blocks:
            day_cell = ws.cell(row, block.time_col, day.value)
            _apply_fill(day_cell, WEEKDAY_FILL, WEEKDAY_FONT, CENTER)
            if block.group_end > block.group_start:
                ws.merge_cells(
                    start_row=row,
                    start_column=block.group_start,
                    end_row=row,
                    end_column=block.group_end,
                )
            for c in range(block.group_start, block.group_end + 1):
                _apply_fill(ws.cell(row, c), WEEKDAY_FILL, None, CENTER)
        row += 1

        for slot in slots:
            block_start = row
            row_heights = [GROUPS_EVENT_ROW_HEIGHT_MIN, GROUPS_EVENT_ROW_HEIGHT_MIN, GROUPS_EVENT_ROW_HEIGHT_MIN]

            for block in blocks:
                ws.cell(block_start, block.time_col, _slot_label(slot.start_time, slot.end_time))
                ws.merge_cells(
                    start_row=block_start,
                    start_column=block.time_col,
                    end_row=block_start + 2,
                    end_column=block.time_col,
                )
                _style_triple_block(
                    ws,
                    start_row=block_start,
                    start_col=block.time_col,
                    end_col=block.time_col,
                    fill=WHITE_FILL,
                    fonts=(NORMAL_FONT, NORMAL_FONT, NORMAL_FONT),
                )

                i = 0
                while i < len(block.columns):
                    col = block.columns[i]
                    meetings = cells.get((day, slot.start_time, col.group_id), [])
                    sign = cell_signature(meetings)
                    span = 1
                    while (
                        i + span < len(block.columns)
                        and cell_signature(cells.get((day, slot.start_time, block.columns[i + span].group_id), []))
                        == sign
                        and sign
                    ):
                        span += 1
                    start_col = block.group_start + i
                    end_col = start_col + span - 1
                    title, instructors, room = _write_event_block(
                        ws,
                        start_row=block_start,
                        start_col=start_col,
                        end_col=end_col,
                        meetings=meetings,
                        instructor_labels=instructor_labels,
                    )
                    if meetings:
                        row_heights[0] = max(row_heights[0], _row_height_for_text(title, span))
                        row_heights[1] = max(row_heights[1], _row_height_for_text(instructors, span))
                        row_heights[2] = max(row_heights[2], _row_height_for_text(room, span))
                    i += span

            for offset, height in enumerate(row_heights):
                ws.row_dimensions[block_start + offset].height = height

            row = block_start + 3

    ws.freeze_panes = "A3"


def _write_compact_groups_sheet(
    ws: Worksheet,
    *,
    columns: list[ExportColumn],
    term: TermConfig,
    occupied_slots: list[tuple[Weekday, TermTimeSlot]],
    cells: dict[tuple[Weekday, dtm.time, str], list[ExportMeeting]],
    instructor_labels: dict[str, str],
    group_sizes: dict[str, int | None],
) -> None:
    time_header = ws.cell(1, 1)
    _apply_fill(time_header, TIME_HEADER_FILL, HEADER_FONT, CENTER)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    _apply_fill(ws.cell(2, 1), TIME_HEADER_FILL, HEADER_FONT, CENTER)
    ws.column_dimensions["A"].width = COMPACT_GROUPS_TIME_COL_WIDTH

    index = 0
    while index < len(columns):
        end_index = index + 1
        while end_index < len(columns) and columns[end_index].year_label == columns[index].year_label:
            end_index += 1
        group_start = index + 2
        group_end = end_index + 1
        year_cell = ws.cell(1, group_start, columns[index].year_label)
        _apply_fill(year_cell, HEADER_FILL, HEADER_FONT, CENTER)
        if group_end > group_start:
            ws.merge_cells(
                start_row=1,
                start_column=group_start,
                end_row=1,
                end_column=group_end,
            )
            for col in range(group_start, group_end + 1):
                _apply_fill(ws.cell(1, col), HEADER_FILL, HEADER_FONT, CENTER)

        for offset, column in enumerate(columns[index:end_index]):
            label = column.group_label
            cell = ws.cell(2, group_start + offset, label)
            _apply_fill(cell, HEADER_FILL, HEADER_FONT, CENTER)
            ws.column_dimensions[get_column_letter(group_start + offset)].width = COMPACT_GROUPS_COL_WIDTH
        index = end_index

    for row, (day, slot) in enumerate(occupied_slots, start=3):
        time_label = f"{COMPACT_WEEKDAY_LABELS[day]}\n{_slot_label(slot.start_time, slot.end_time)}"
        time_cell = ws.cell(row, 1, time_label)
        _apply_fill(time_cell, WHITE_FILL, NORMAL_FONT, CENTER)

        row_height = COMPACT_GROUPS_ROW_HEIGHT_MIN
        index = 0
        while index < len(columns):
            column = columns[index]
            meetings = cells.get((day, slot.start_time, column.group_id), [])
            signature = cell_signature(meetings)
            span = 1
            while (
                index + span < len(columns)
                and columns[index + span].year_label == column.year_label
                and cell_signature(cells.get((day, slot.start_time, columns[index + span].group_id), [])) == signature
                and signature
            ):
                span += 1
            start_col = index + 2
            end_col = start_col + span - 1

            if meetings:
                meeting = meetings[0]
                title = meeting.course_short_name or meeting.course
                instructors = format_instructors(meeting.instructors, instructor_labels)
                details = [part for part in (instructors, meeting.room) if part]
                text = "\n".join((title, *details))
                fill = PatternFill("solid", fgColor=course_fill_color(meeting.course))
                rich_text = CellRichText(
                    TextBlock(COMPACT_TITLE_INLINE_FONT, title),
                    *(TextBlock(COMPACT_DETAIL_INLINE_FONT, f"\n{detail}") for detail in details),
                )
                ws.cell(row, start_col, rich_text)
                if end_col > start_col:
                    ws.merge_cells(
                        start_row=row,
                        start_column=start_col,
                        end_row=row,
                        end_column=end_col,
                    )
                for col in range(start_col, end_col + 1):
                    _apply_fill(ws.cell(row, col), fill, None, CENTER)
                width = COMPACT_GROUPS_COL_WIDTH * span
                row_height = max(
                    row_height,
                    _wrapped_line_count(text, width) * COMPACT_GROUPS_LINE_HEIGHT,
                )
            else:
                _apply_fill(ws.cell(row, start_col), WHITE_FILL, None, CENTER)
            index += span

        ws.row_dimensions[row].height = row_height

    ws.freeze_panes = "B3"


def _term_weeks(term: TermConfig) -> list[tuple[dtm.date, dtm.date]]:
    """List (week_start, week_end) covering the union semester window, clipped to bounds."""
    bounds = union_semester_window(term)
    start = bounds.start_date
    end = bounds.end_date
    weeks: list[tuple[dtm.date, dtm.date]] = []
    cursor = week_start_for_date(start, term.starting_day)
    while cursor <= end:
        week_end = cursor + dtm.timedelta(days=6)
        weeks.append((max(cursor, start), min(week_end, end)))
        cursor += dtm.timedelta(days=7)
    return weeks


def _compact_meeting_label(meeting: ExportMeeting) -> str:
    name = (meeting.course_short_name or meeting.course).strip() or "—"
    tag = meeting.tag.strip().lower()
    if tag and tag not in name.casefold() and tag not in {"class"}:
        name = f"{name} ({tag})" if "(" not in name else name
    room = meeting.room.strip()
    main = f"{name} {room}".strip() if room else name
    groups = list(meeting.groups)
    short = (meeting.course_short_name or "").strip().casefold()
    if len(groups) == 1 and short and groups[0].casefold() == short:
        return main
    if groups and not (len(groups) == 1 and groups[0].casefold() == name.casefold()):
        shown = groups[:3]
        suffix = ", ".join(shown) + (", ..." if len(groups) > 3 else "")
        return f"{main} ({suffix})"
    return main


def _course_audience_tokens(course: CourseConfig) -> list[str]:
    tokens: list[str] = []
    seen: set[str] = set()
    for component in course.components:
        for token in component.audience:
            if token in seen:
                continue
            seen.add(token)
            tokens.append(token)
        for series in component.sessions or []:
            for token in series.audience or []:
                if token in seen:
                    continue
                seen.add(token)
                tokens.append(token)
    return tokens


def build_calendar_course_legend(
    *,
    courses: list[CourseConfig],
    meetings: list[ExportMeeting],
    groups: set[str],
    instructor_labels: dict[str, str],
    selector_map: dict[str, set[str]],
    section_code: str | None = None,
) -> list[tuple[str, str, str]]:
    if not groups:
        return []

    instructors_by_course: dict[str, set[str]] = defaultdict(set)
    for meeting in meetings:
        if not any(group in groups for group in meeting.groups):
            continue
        course_name = meeting.course.strip()
        if not course_name:
            continue
        for instructor_id in meeting.instructors:
            label = instructor_labels.get(instructor_id, instructor_id).strip()
            if label:
                instructors_by_course[course_name].add(label)

    rows: list[tuple[str, str, str]] = []
    for course in courses:
        if section_code is not None and (course.section_code or "").strip() != section_code:
            continue
        tokens = _course_audience_tokens(course)
        expanded = expand_group_tokens(tokens, selector_map)
        if not any(group in groups for group in expanded):
            continue
        course_name = course.name.strip()
        if not course_name:
            continue
        course_instructors = [
            instructor_labels.get(item.id, item.id).strip()
            for item in (course.instructors or [])
            if (instructor_labels.get(item.id, item.id) or "").strip()
        ]
        meeting_instructors = sorted(instructors_by_course.get(course_name, set()))
        instructor = ", ".join(course_instructors or meeting_instructors) or "—"
        short_name = (course.short_name or "").strip() or course_name
        rows.append((short_name, course_name, instructor))

    rows.sort(key=lambda row: row[0].casefold())
    return rows


def _write_calendar_course_legend(
    ws: Worksheet,
    *,
    start_col: int,
    rows: list[tuple[str, str, str]],
) -> None:
    for offset, header in enumerate(CALENDAR_LEGEND_HEADERS):
        cell = ws.cell(1, start_col + offset, header)
        _apply_fill(cell, HEADER_FILL, HEADER_FONT, CENTER)
        ws.column_dimensions[get_column_letter(start_col + offset)].width = CALENDAR_LEGEND_COL_WIDTHS[offset]

    for row_index, values in enumerate(rows, start=2):
        for offset, value in enumerate(values):
            cell = ws.cell(row_index, start_col + offset, value)
            _apply_fill(cell, WHITE_FILL, NORMAL_FONT, LEFT)


def _write_calendar_sheet(
    ws: Worksheet,
    *,
    term: TermConfig,
    meetings: list[ExportMeeting],
    legend_rows: list[tuple[str, str, str]] | None = None,
) -> None:
    days = _term_days(term)
    slots = list(term.time_slots)
    weeks = _term_weeks(term)

    # Index meetings by date + slot start
    slot_starts = {slot.start_time for slot in slots}
    by_key: dict[tuple[dtm.date, dtm.time], list[ExportMeeting]] = defaultdict(list)
    for meeting in meetings:
        row_start = meeting.start if meeting.start in slot_starts else _nearest_slot_start(meeting.start, slots)
        if row_start is None:
            continue
        by_key[(meeting.date, row_start)].append(meeting)

    # Header: weekdays
    for i, day in enumerate(days):
        cell = ws.cell(1, i + 2, day.value.capitalize())
        _apply_fill(cell, WHITE_FILL, TITLE_FONT, CENTER)
    _apply_fill(ws.cell(1, 1), WHITE_FILL)

    row = 2
    for week_number, (week_start, _week_end) in enumerate(weeks, start=1):
        week_cell = ws.cell(row, 1, f"Week {week_number}")
        _apply_fill(week_cell, WEEK_HEADER_FILL, TITLE_FONT, CENTER)
        for i, day in enumerate(days):
            day_date = week_start_for_date(week_start, term.starting_day) + dtm.timedelta(days=day.index)
            label = day_date.strftime("%B ") + str(day_date.day)
            cell = ws.cell(row, i + 2, label)
            _apply_fill(cell, WEEK_HEADER_FILL, TITLE_FONT, CENTER)
        row += 1

        for slot in slots:
            time_cell = ws.cell(row, 1, _slot_label(slot.start_time, slot.end_time))
            _apply_fill(time_cell, WHITE_FILL, NORMAL_FONT, CENTER)
            for i, day in enumerate(days):
                day_date = week_start_for_date(week_start, term.starting_day) + dtm.timedelta(days=day.index)
                cell_meetings = by_key.get((day_date, slot.start_time), [])
                unique: dict[str, ExportMeeting] = {}
                for meeting in cell_meetings:
                    unique.setdefault(pattern_signature(meeting), meeting)
                text = "\n".join(_compact_meeting_label(m) for m in unique.values())
                cell = ws.cell(row, i + 2, text or None)
                _apply_fill(cell, WHITE_FILL, NORMAL_FONT, LEFT)
            row += 1

        if week_number < len(weeks):
            row += 1

    ws.column_dimensions["A"].width = 14
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 18
    ws.freeze_panes = "B2"

    if legend_rows is not None:
        _write_calendar_course_legend(ws, start_col=len(days) + 3, rows=legend_rows)


def _write_table_sheet(
    ws: Worksheet,
    *,
    headers: tuple[str, ...],
    col_widths: tuple[int, ...],
    rows: list[tuple[object, ...]],
) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        _apply_fill(cell, HEADER_FILL, HEADER_FONT, CENTER)
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

    for row_index, values in enumerate(rows, start=2):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            _apply_fill(cell, WHITE_FILL, NORMAL_FONT, LEFT)

    ws.freeze_panes = "A2"
    last_row = max(1, 1 + len(rows))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


def _track_group_maps(sections: list[SectionConfig]) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    """Return ``group→tracks`` and ``track→groups`` maps (``Program/Track`` labels)."""
    group_to_tracks: dict[str, set[str]] = defaultdict(set)
    track_to_groups: dict[str, set[str]] = {}
    for section in sections:
        for program in section.programs:
            for track in program.tracks:
                if not track.groups:
                    continue
                label = f"{program.code}/{track.name}"
                groups = set(track.groups)
                track_to_groups[label] = groups
                for group_code in groups:
                    group_to_tracks[group_code].add(label)
    return group_to_tracks, track_to_groups


def _course_student_group_tokens(course: CourseConfig) -> list[str]:
    tokens: list[str] = []
    seen: set[str] = set()
    for component in course.components:
        for token in component.audience:
            if token in seen:
                continue
            seen.add(token)
            tokens.append(token)
    return tokens


def _format_course_groups_by_track(
    course: CourseConfig,
    *,
    selector_map: dict[str, set[str]],
    group_to_tracks: dict[str, set[str]],
    track_to_groups: dict[str, set[str]],
    section_groups: set[str] | None = None,
) -> str:
    """Groups column: track-grouped only when every group of that track is present."""
    tokens = _course_student_group_tokens(course)
    groups = expand_group_tokens(tokens, selector_map)
    if section_groups is not None:
        groups &= section_groups
    if not groups:
        return ""

    candidate_tracks: set[str] = set()
    for group_code in groups:
        candidate_tracks.update(group_to_tracks.get(group_code, ()))

    remaining = set(groups)
    lines: list[str] = []
    for track in sorted(candidate_tracks, key=str.casefold):
        track_groups = track_to_groups.get(track, set())
        if not track_groups or not track_groups <= groups:
            continue
        lines.append(f"{track}: {', '.join(sorted(track_groups, key=str.casefold))}")
        remaining -= track_groups

    if remaining:
        lines.append(", ".join(sorted(remaining, key=str.casefold)))
    return "\n".join(lines)


def _course_instructor_role_labels(
    course: CourseConfig,
    instructor_labels: dict[str, str],
) -> str:
    parts: list[str] = []
    for assignment in course.instructors:
        name = instructor_labels.get(assignment.id, assignment.id)
        role = (assignment.role or "").strip()
        parts.append(f"{name} ({role})" if role else name)
    return "\n".join(parts)


def _instructor_courses_column(
    instructor_id: str,
    courses: list[CourseConfig],
) -> str:
    parts: list[str] = []
    for course in courses:
        for assignment in course.instructors:
            if assignment.id != instructor_id:
                continue
            role = (assignment.role or "").strip()
            label = f"{course.name} ({role})" if role else course.name
            parts.append(label)
    return "\n".join(parts)


def _write_distributions_sheet(ws: Worksheet, config: ScheduleConfig) -> None:
    """One row per student–group membership (import-friendly roster)."""
    by_code = {group.code: group for group in config.students_groups}
    rows: list[tuple[object, ...]] = []
    for section in config.term.sections:
        section_label = (section.name or section.code).strip() or section.code
        for code in iter_section_group_codes(section):
            group = by_code.get(code)
            if group is None:
                continue
            group_label = (group.name or group.code).strip() or group.code
            for email in sorted(group.students, key=str.casefold):
                rows.append((email, group_label, section_label))
    _write_table_sheet(
        ws,
        headers=DISTRIBUTIONS_HEADERS,
        col_widths=DISTRIBUTIONS_COL_WIDTHS,
        rows=rows,
    )


def _write_instructors_sheet(ws: Worksheet, config: ScheduleConfig) -> None:
    rows: list[tuple[object, ...]] = []
    for instructor in config.instructors:
        rows.append(
            (
                instructor.name_en or "",
                instructor.name_ru or "",
                instructor.email or "",
                instructor.alias or "",
                instructor.position or "",
                _instructor_courses_column(instructor.id, config.courses),
                instructor.id,
            )
        )
    _write_table_sheet(
        ws,
        headers=INSTRUCTORS_HEADERS,
        col_widths=INSTRUCTORS_COL_WIDTHS,
        rows=rows,
    )


def _write_subjects_sheet(
    ws: Worksheet,
    config: ScheduleConfig,
    instructor_labels: dict[str, str],
) -> None:
    selector_map = build_selector_map(
        SectionsConfig(sections=config.term.sections, students_groups=config.students_groups)
    )
    rows: list[tuple[object, ...]] = []
    matched_course_names: set[str] = set()

    for section in config.term.sections:
        section_label = (section.name or section.code).strip() or section.code
        section_groups = section_group_set(section)
        group_to_tracks, track_to_groups = _track_group_maps([section])
        for course in config.courses:
            groups = _format_course_groups_by_track(
                course,
                selector_map=selector_map,
                group_to_tracks=group_to_tracks,
                track_to_groups=track_to_groups,
                section_groups=section_groups,
            )
            if not groups:
                continue
            matched_course_names.add(course.name)
            rows.append(
                (
                    section_label,
                    course.short_name or "",
                    course.name,
                    course.name_ru or "",
                    groups,
                    _course_instructor_role_labels(course, instructor_labels),
                )
            )

    group_to_tracks, track_to_groups = _track_group_maps(config.term.sections)
    for course in config.courses:
        if course.name in matched_course_names:
            continue
        groups = _format_course_groups_by_track(
            course,
            selector_map=selector_map,
            group_to_tracks=group_to_tracks,
            track_to_groups=track_to_groups,
        )
        rows.append(
            (
                "",
                course.short_name or "",
                course.name,
                course.name_ru or "",
                groups,
                _course_instructor_role_labels(course, instructor_labels),
            )
        )

    col_count = len(SUBJECTS_HEADERS)
    for col, header in enumerate(SUBJECTS_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        _apply_fill(cell, HEADER_FILL, HEADER_FONT, CENTER)
        ws.column_dimensions[get_column_letter(col)].width = SUBJECTS_COL_WIDTHS[col - 1]

    # Contiguous rows sharing the Section value get one outer border.
    block_ranges: list[tuple[int, int]] = []
    index = 0
    while index < len(rows):
        section_key = rows[index][0]
        end = index + 1
        while end < len(rows) and rows[end][0] == section_key:
            end += 1
        block_ranges.append((index, end - 1))
        index = end

    for block_start, block_end in block_ranges:
        start_row = block_start + 2
        end_row = block_end + 2
        for row_offset in range(block_start, block_end + 1):
            excel_row = row_offset + 2
            values = rows[row_offset]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(excel_row, col, value)
                _apply_fill(
                    cell,
                    WHITE_FILL,
                    NORMAL_FONT,
                    LEFT,
                    border=_section_block_border(
                        row=excel_row,
                        col=col,
                        start_row=start_row,
                        end_row=end_row,
                        start_col=1,
                        end_col=col_count,
                    ),
                )

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    last_row = max(1, 1 + len(rows))
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{last_row}"


def _append_summary_sheets(
    wb: Workbook,
    config: ScheduleConfig,
    *,
    used_names: set[str],
    instructor_labels: dict[str, str],
) -> None:
    distributions_title = _unique_sheet_name(DISTRIBUTIONS_SHEET_NAME, used_names)
    used_names.add(distributions_title)
    distributions = wb.create_sheet(distributions_title)
    _write_distributions_sheet(distributions, config)

    instructors_title = _unique_sheet_name(INSTRUCTORS_SHEET_NAME, used_names)
    used_names.add(instructors_title)
    instructors = wb.create_sheet(instructors_title)
    _write_instructors_sheet(instructors, config)

    subjects_title = _unique_sheet_name(SUBJECTS_SHEET_NAME, used_names)
    used_names.add(subjects_title)
    subjects = wb.create_sheet(subjects_title)
    _write_subjects_sheet(subjects, config, instructor_labels)


def _write_section_sheet(
    ws: Worksheet,
    *,
    config: ScheduleConfig,
    section: SectionConfig,
    all_meetings: list[ExportMeeting],
    instructor_labels: dict[str, str],
    group_sizes: dict[str, int | None],
) -> None:
    meetings = filter_meetings_for_section(all_meetings, section)
    layout = section_export_layout(section)
    if layout in {"groups", "compact_groups"}:
        columns = build_columns(
            section,
            meetings,
            config,
            include_unused=layout == "compact_groups",
        )
        cells = _collapse_weekly_cells(meetings, config.term.time_slots)
        if layout == "compact_groups":
            _write_compact_groups_sheet(
                ws,
                columns=columns,
                term=config.term,
                occupied_slots=_compact_occupied_slots(meetings, config.term),
                cells=cells,
                instructor_labels=instructor_labels,
                group_sizes=group_sizes,
            )
            return
        _write_groups_sheet(
            ws,
            columns=columns,
            term=config.term,
            cells=cells,
            instructor_labels=instructor_labels,
            group_sizes=group_sizes,
        )
        return
    selector_map = build_selector_map(
        SectionsConfig(sections=config.term.sections, students_groups=config.students_groups)
    )
    legend_rows = build_calendar_course_legend(
        courses=config.courses,
        meetings=meetings,
        groups=section_group_set(section),
        instructor_labels=instructor_labels,
        selector_map=selector_map,
        section_code=section.code,
    )
    _write_calendar_sheet(ws, term=config.term, meetings=meetings, legend_rows=legend_rows)


def _next_sheet(wb: Workbook, title: str, *, first: bool) -> Worksheet:
    if first:
        ws = wb.active
        assert ws is not None
        ws.title = title
        return ws
    return wb.create_sheet(title)


def build_export_workbook(config: ScheduleConfig) -> Workbook:
    all_meetings = expand_meetings(config)
    instructor_labels = _instructor_label_by_id(config)
    group_sizes = _group_sizes(config)
    selector_map = build_selector_map(
        SectionsConfig(sections=config.term.sections, students_groups=config.students_groups)
    )
    wb = Workbook()
    used_names: set[str] = set()
    sections = list(config.term.sections)
    first_sheet = True

    if not sections:
        ws = wb.active
        assert ws is not None
        ws.title = "Schedule"
        used_names.add("Schedule")
        _append_summary_sheets(wb, config, used_names=used_names, instructor_labels=instructor_labels)
        return wb

    for section in sections:
        layout = section_export_layout(section)
        if layout == "calendar" and len(section.programs) > 1:
            for program in section.programs:
                section_name = (section.name or section.code).strip()
                program_name = (program.name or program.code).strip()
                title = _unique_sheet_name(
                    f"{section_name} {program_name}",
                    used_names,
                )
                used_names.add(title)
                ws = _next_sheet(wb, title, first=first_sheet)
                first_sheet = False
                program_groups = program_group_set(program)
                program_meetings = filter_meetings_for_groups(all_meetings, program_groups)
                legend_rows = build_calendar_course_legend(
                    courses=config.courses,
                    meetings=program_meetings,
                    groups=program_groups,
                    instructor_labels=instructor_labels,
                    selector_map=selector_map,
                    section_code=section.code,
                )
                _write_calendar_sheet(
                    ws,
                    term=config.term,
                    meetings=program_meetings,
                    legend_rows=legend_rows,
                )
            continue

        title = _unique_sheet_name(section.name or section.code, used_names)
        used_names.add(title)
        ws = _next_sheet(wb, title, first=first_sheet)
        first_sheet = False
        _write_section_sheet(
            ws,
            config=config,
            section=section,
            all_meetings=all_meetings,
            instructor_labels=instructor_labels,
            group_sizes=group_sizes,
        )

    _append_summary_sheets(wb, config, used_names=used_names, instructor_labels=instructor_labels)
    return wb


def export_schedule_xlsx(config: ScheduleConfig) -> tuple[bytes, str]:
    wb = build_export_workbook(config)
    buffer = io.BytesIO()
    wb.save(buffer)
    term_label = (config.term.name or "").strip() or "Schedule"
    return buffer.getvalue(), export_filename(term_label)
