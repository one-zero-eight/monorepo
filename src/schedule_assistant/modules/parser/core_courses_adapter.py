import datetime as dtm
from collections import defaultdict
from collections.abc import Generator

from openpyxl.utils import coordinate_to_tuple, get_column_letter

from src.logging_ import logger
from src.schedule_assistant.core_courses.cell_to_event import CoreCourseEvent
from src.schedule_assistant.core_courses.config import CoreCoursesConfig, Target
from src.schedule_assistant.core_courses.location_parser import Item
from src.schedule_assistant.core_courses.parser import CoreCoursesParser, use
from src.schedule_assistant.utils import (
    WEEKDAYS,
    fetch_xlsx_spreadsheet,
    get_sheet_gids,
    nearest_weekday,
    sanitize_sheet_name,
)

from .schemas import CoreCourseComponent, GroupedCoreCourse, Lesson


async def _collect_core_courses_lessons(parser_config: CoreCoursesConfig) -> list[Lesson]:
    parser = CoreCoursesParser()
    xlsx_file = await fetch_xlsx_spreadsheet(spreadsheet_id=parser_config.spreadsheet_id)
    original_target_sheet_names = [target.sheet_name for target in parser_config.targets]
    sheet_gids = await get_sheet_gids(parser_config.spreadsheet_id)
    pipeline_result = list(
        parser.pipeline(xlsx_file, original_target_sheet_names, sheet_gids, parser_config.spreadsheet_id)
    )
    dfs_merged_ranges = parser.last_dfs_merged_ranges
    assert dfs_merged_ranges is not None

    all_lessons: list[Lesson] = []
    for target, grouped_dfs_with_cells_list in zip(parser_config.targets, pipeline_result):
        merged_ranges = dfs_merged_ranges.get(sanitize_sheet_name(target.sheet_name))

        # merge range index -> list of events
        merged_registry_for_events: dict[int, list[CoreCourseEvent]] = defaultdict(list)
        non_merged_events: list[CoreCourseEvent] = []

        for grouped_dfs_with_cells in grouped_dfs_with_cells_list:
            series_with_generators = grouped_dfs_with_cells.apply(
                use,
                target=target,
                dont_care_location_string=parser_config.dont_care_location_string,
            )
            for generator in series_with_generators:
                generator: Generator[CoreCourseEvent]

                for cell_event in generator:
                    if cell_event.subject in parser_config.ignored_subjects:
                        logger.debug(f"> Ignoring {cell_event.subject}")
                        continue

                    if merged_ranges:
                        if not cell_event.a1:
                            non_merged_events.append(cell_event)
                            continue
                        cell_row, cell_col = coordinate_to_tuple(cell_event.a1)
                        cell_row -= 1
                        cell_col -= 1
                        for i, (min_row, min_col, max_row, max_col) in enumerate(merged_ranges):
                            if (min_row <= cell_row <= max_row) and (min_col <= cell_col <= max_col):
                                merged_registry_for_events[i].append(cell_event)
                                break
                        else:
                            non_merged_events.append(cell_event)
                    else:
                        non_merged_events.append(cell_event)

        lessons_from_non_merged: list[Lesson] = []
        for cell_event in non_merged_events:
            if cell_event.location_item is None:
                lessons_from_non_merged.append(_event_to_lesson(cell_event))
            else:
                lessons_from_non_merged.extend(_process_location_item(cell_event, target))

        lessons_from_merged: list[Lesson] = []
        for merged_range_index, _events in merged_registry_for_events.items():
            assert merged_ranges is not None
            _events: list[CoreCourseEvent]
            groups = []
            merged_students_number = []
            merged_range = merged_ranges[merged_range_index]
            min_row, min_col, max_row, max_col = merged_range
            merged_range_a1_range = (
                f"{get_column_letter(min_col + 1)}{min_row + 1}:{get_column_letter(max_col + 1)}{max_row + 1}"
            )
            excel_ranges = []
            cell_event = _events[0]

            for cell_event in _events:
                if cell_event.group:
                    if cell_event.group not in groups and cell_event.group_student_number:
                        merged_students_number.append(cell_event.group_student_number)
                    if cell_event.group not in groups and cell_event.a1:
                        excel_ranges.append(cell_event.a1)
                else:
                    if cell_event.group_student_number:
                        merged_students_number.append(cell_event.group_student_number)
                    if cell_event.a1:
                        excel_ranges.append(cell_event.a1)

                if isinstance(cell_event.group, list | tuple):
                    groups.extend(cell_event.group)
                elif cell_event.group:
                    groups.append(cell_event.group)

            if cell_event.location_item is None:
                lessons_from_merged.append(
                    _event_to_lesson(
                        cell_event,
                        group_name=tuple(groups),
                        students_number=sum(merged_students_number) if merged_students_number else None,
                        a1_range=merged_range_a1_range,
                    )
                )
            else:
                lessons_from_merged.extend(
                    _process_location_item(
                        cell_event,
                        target,
                        group_name=tuple(groups),
                        students_number=sum(merged_students_number) if merged_students_number else None,
                        a1_range=merged_range_a1_range,
                    )
                )
        logger.info(
            f"For {target.sheet_name} found {len(lessons_from_non_merged)} non-merged lessons and {len(lessons_from_merged)} merged (same a1 range) lessons, totaling {len(lessons_from_merged + lessons_from_non_merged)} lessons"
        )
        merged_lessons = merge_identical_lessons(lessons_from_merged + lessons_from_non_merged)
        logger.info(f"After merging identical lessons, for {target.sheet_name} found {len(merged_lessons)} lessons")
        all_lessons.extend(merged_lessons)

    def _sort_key(x: Lesson):
        group = x.group_name
        if group is None:
            group = ()
        elif isinstance(group, str):
            group = (group,)
        return (x.course_name or "", group, x.weekday or "", x.start_time)

    all_lessons.sort(key=_sort_key)
    return all_lessons


async def get_grouped_core_courses_lessons(parser_config: CoreCoursesConfig) -> list[GroupedCoreCourse]:
    flat_lessons = await _collect_core_courses_lessons(parser_config)
    term_dates = {target.sheet_name: (target.start_date, target.end_date) for target in parser_config.targets}
    grouped = group_core_course_lessons(flat_lessons, term_dates=term_dates)
    logger.info(f"Grouped core course lessons: {len(flat_lessons)} -> {len(grouped)}")
    return grouped


def _modifiers_for_output(item: Item | None) -> Item | None:
    """Omit modifiers when the location string is just a room (no temporal/nested rules)."""
    if item is None:
        return None
    has_extra = bool(
        item.starts_from
        or item.ends_on
        or item.starts_at
        or item.till
        or item.on_weeks
        or item.on
        or item.except_
        or item.NEST
    )
    if not has_extra:
        return None
    return item


def _item_for_lesson_dates(lesson: Lesson) -> Item | None:
    """Build modifier Item for a component, including per-row date overrides from parsing."""
    if lesson.modifiers is not None:
        item = lesson.modifiers.model_copy(deep=True)
        if lesson.date_on is not None:
            item.on = lesson.date_on
        if lesson.date_except is not None:
            item.except_ = lesson.date_except
        if lesson.date_from is not None:
            item.starts_from = lesson.date_from
        return item
    if lesson.date_on or lesson.date_except or lesson.date_from:
        location = lesson.room if isinstance(lesson.room, str) else None
        return Item(
            location=location,
            on=lesson.date_on,
            except_=lesson.date_except,
            starts_from=lesson.date_from,
        )
    return None


def _lesson_dates_from_modifiers(
    item: Item | None,
) -> tuple[list[dtm.date] | None, list[dtm.date] | None, dtm.date | None]:
    if item is None:
        return None, None, None
    return item.on, item.except_, item.starts_from


def _modifiers_for_component(lesson: Lesson) -> Item | None:
    return _modifiers_for_output(_item_for_lesson_dates(lesson))


def _as_audience_list(group_name: str | tuple[str, ...] | list[str] | None) -> list[str]:
    if group_name is None:
        return []
    if isinstance(group_name, str):
        return [group_name]
    return list(group_name)


def _audience_to_group_name(audience: list[str]) -> str | tuple[str, ...] | None:
    if not audience:
        return None
    if len(audience) == 1:
        return audience[0]
    return tuple(audience)


def _component_sort_key(component: CoreCourseComponent) -> tuple:
    group = tuple(component.audience)
    type_rank = {"lec": 0, "лек": 0, "tut": 1, "тут": 1, "lab": 2, "лаб": 2}.get(str(component.type or "").lower(), 99)
    return (type_rank, component.weekday or "", component.start_time, group)


def _lesson_to_component(lesson: Lesson) -> CoreCourseComponent:
    return CoreCourseComponent(
        type=lesson.lesson_class_type,
        weekday=lesson.weekday,
        start_time=lesson.start_time,
        end_time=lesson.end_time,
        room=lesson.room,
        instructor=lesson.teacher,
        audience=_as_audience_list(lesson.group_name),
        students_number=lesson.students_number,
        modifiers=_modifiers_for_component(lesson),
        a1_range=lesson.a1_range,
    )


def grouped_core_course_to_json(course: GroupedCoreCourse) -> dict:
    data = course.model_dump(mode="json")
    for component in data.get("components", []):
        if component.get("modifiers") is None:
            component.pop("modifiers", None)
        for key in ("date_on", "date_except", "date_from"):
            component.pop(key, None)
    return data


def group_core_course_lessons(
    lessons: list[Lesson],
    *,
    term_dates: dict[str, tuple[dtm.date, dtm.date]],
) -> list[GroupedCoreCourse]:
    by_key: dict[tuple, list[Lesson]] = defaultdict(list)
    for lesson in lessons:
        start_date, end_date = term_dates.get(lesson.google_sheet_name, (None, None))
        key = (
            lesson.course_name or "",
            lesson.lesson_name,
            lesson.spreadsheet_id,
            lesson.google_sheet_gid,
            lesson.google_sheet_name,
            start_date,
            end_date,
        )
        by_key[key].append(lesson)

    grouped: list[GroupedCoreCourse] = []
    for (cohort, subject, spreadsheet_id, google_sheet_gid, google_sheet_name, start_date, end_date), bucket in sorted(
        by_key.items(), key=lambda item: (item[0][0], item[0][1], item[0][4])
    ):
        components = sorted((_lesson_to_component(lesson) for lesson in bucket), key=_component_sort_key)
        grouped.append(
            GroupedCoreCourse(
                cohort=cohort or None,
                subject=subject,
                spreadsheet_id=spreadsheet_id,
                google_sheet_gid=google_sheet_gid,
                google_sheet_name=google_sheet_name,
                start_date=start_date,
                end_date=end_date,
                components=components,
            )
        )
    return grouped


def expand_grouped_core_course_lessons(grouped: list[GroupedCoreCourse]) -> list[Lesson]:
    flat: list[Lesson] = []
    for course in grouped:
        for component in course.components:
            date_on, date_except, date_from = _lesson_dates_from_modifiers(component.modifiers)
            flat.append(
                Lesson(
                    lesson_name=course.subject,
                    lesson_class_type=component.type,
                    weekday=component.weekday,
                    start_time=component.start_time,
                    end_time=component.end_time,
                    course_name=course.cohort,
                    group_name=_audience_to_group_name(component.audience),
                    teacher=component.instructor,
                    room=component.room,
                    source_type="core_course",
                    date_on=date_on,
                    date_except=date_except,
                    date_from=date_from,
                    modifiers=component.modifiers,
                    students_number=component.students_number,
                    spreadsheet_id=course.spreadsheet_id,
                    google_sheet_gid=course.google_sheet_gid,
                    google_sheet_name=course.google_sheet_name,
                    a1_range=component.a1_range,
                )
            )

    def _sort_key(x: Lesson):
        group = x.group_name
        if group is None:
            group = ()
        elif isinstance(group, str):
            group = (group,)
        return (x.course_name or "", group, x.weekday or "", x.start_time)

    flat.sort(key=_sort_key)
    return flat


def _event_to_lesson(
    cell_event: CoreCourseEvent,
    *,
    group_name: str | tuple[str, ...] | None = None,
    students_number: int | None = None,
    a1_range: str | None = None,
) -> Lesson:
    return Lesson(
        lesson_name=cell_event.subject,
        lesson_class_type=cell_event.class_type,
        weekday=WEEKDAYS[cell_event.weekday],
        start_time=cell_event.start_time,
        end_time=cell_event.end_time,
        course_name=cell_event.course,
        group_name=group_name if group_name is not None else cell_event.group,
        teacher=cell_event.teacher,
        room=cell_event.location,
        source_type="core_course",
        date_on=None,
        date_except=None,
        students_number=students_number if students_number is not None else cell_event.group_student_number,
        spreadsheet_id=cell_event.spreadsheet_id,
        google_sheet_gid=cell_event.google_sheet_gid,
        google_sheet_name=cell_event.google_sheet_name,
        a1_range=a1_range if a1_range is not None else cell_event.a1,
        modifiers=cell_event.location_item,
    )


def _expand_nested_location_lessons(lessons: list[Lesson]) -> list[Lesson]:
    """Expand lessons with nested location modifiers for collision checks (flat ONLY ON rows)."""
    result: list[Lesson] = []
    for lesson in lessons:
        item = lesson.modifiers
        if not item or not item.NEST:
            result.append(lesson)
            continue
        nested_on = [nested for nested in item.NEST if nested.on]
        if not nested_on:
            result.append(lesson)
            continue

        main = lesson.model_copy()
        main.date_except = list(main.date_except or [])
        for nested in nested_on:
            if nested.location and nested.on:
                main.date_except = sorted(set(main.date_except) | set(nested.on))
        if not main.date_except:
            main.date_except = None
        result.append(main)

        for nested in nested_on:
            override = lesson.model_copy()
            override.date_on = nested.on
            override.room = nested.location or lesson.room
            override.start_time = nested.starts_at or lesson.start_time
            override.end_time = nested.till or lesson.end_time
            result.append(override)
    return result


def _process_location_item(
    cell_event: CoreCourseEvent,
    target: Target,
    *,
    group_name: str | tuple[str, ...] | None = None,
    students_number: int | None = None,
    a1_range: str | None = None,
) -> list[Lesson]:
    location_item = cell_event.location_item
    assert location_item is not None
    starts = location_item.starts_from or target.start_date

    def convert_weeks_on_to_only_on(item: Item):
        if item.on_weeks:
            on = []
            for week in item.on_weeks:
                on_date = nearest_weekday(starts, cell_event.weekday) + dtm.timedelta(weeks=week - 1)
                on.append(on_date)
            if item.on:
                item.on.extend(on)
            elif on:
                item.on = on
        if item.on:
            item.on = sorted(set(item.on))

    lesson_start_time = cell_event.start_time
    lesson_end_time = cell_event.end_time

    if location_item.starts_at:
        _start_time = dtm.datetime.combine(starts, cell_event.start_time)
        _end_time = dtm.datetime.combine(starts, cell_event.end_time)
        duration = _end_time - _start_time
        lesson_start_time = location_item.starts_at
        lesson_end_time = (dtm.datetime.combine(starts, lesson_start_time) + duration).time()
    if location_item.till:
        lesson_end_time = location_item.till

    convert_weeks_on_to_only_on(location_item)

    def build_lesson() -> Lesson:
        return Lesson(
            lesson_name=cell_event.subject,
            lesson_class_type=cell_event.class_type,
            weekday=WEEKDAYS[cell_event.weekday],
            start_time=lesson_start_time,
            end_time=lesson_end_time,
            course_name=cell_event.course,
            group_name=group_name if group_name is not None else cell_event.group,
            teacher=cell_event.teacher,
            room=location_item.location or cell_event.location,
            date_on=location_item.on,
            date_except=location_item.except_,
            date_from=location_item.starts_from,
            students_number=students_number if students_number is not None else cell_event.group_student_number,
            spreadsheet_id=cell_event.spreadsheet_id,
            google_sheet_gid=cell_event.google_sheet_gid,
            google_sheet_name=cell_event.google_sheet_name,
            a1_range=a1_range if a1_range is not None else cell_event.a1,
            modifiers=cell_event.location_item,
        )

    if location_item.NEST:
        extra_nested: list[Item] = []
        for nested in location_item.NEST:
            convert_weeks_on_to_only_on(nested)
            if not nested.on:
                logger.info(f"Root Item: {location_item}, {nested}")
                extra_nested.append(nested)
        if extra_nested:  # TODO: Handle '421 (316 FROM 31/10)' case
            logger.warning(f"Extra nested is not implemented yet\nItem({location_item})")
        return [build_lesson()]

    return [build_lesson()]


def _are_lessons_identical(lesson1: Lesson, lesson2: Lesson) -> bool:
    """Check if two lessons are identical (excluding Excel cell location)"""
    return (
        lesson1.lesson_name == lesson2.lesson_name
        and lesson1.weekday == lesson2.weekday
        and lesson1.start_time == lesson2.start_time
        and lesson1.end_time == lesson2.end_time
        and lesson1.room == lesson2.room
        and lesson1.teacher == lesson2.teacher
        and lesson1.date_on == lesson2.date_on
        and lesson1.date_except == lesson2.date_except
    )


def merge_identical_lessons(lessons: list[Lesson]) -> list[Lesson]:
    groups: list[list[Lesson]] = []
    for lesson in lessons:
        # Find if this lesson belongs to an existing group
        found_group = False
        for group in groups:
            if _are_lessons_identical(lesson, group[0]):
                group.append(lesson)
                found_group = True
                break
        if not found_group:
            groups.append([lesson])

    result = []
    for group in groups:
        if len(group) > 1:
            excel_ranges = [lesson.a1_range for lesson in group if lesson.a1_range is not None]
            merged_groups = []
            for lesson in group:
                if lesson.group_name is not None:
                    if isinstance(lesson.group_name, tuple):
                        merged_groups.extend(lesson.group_name)
                    else:
                        merged_groups.append(lesson.group_name)
            students_numbers = [lesson.students_number for lesson in group if lesson.students_number is not None]
            students_number = sum(students_numbers) if students_numbers else None
            lesson = group[0].model_copy()
            lesson.a1_range = ";".join(excel_ranges)
            lesson.group_name = tuple(sorted(merged_groups))
            lesson.students_number = students_number
            result.append(lesson)
        else:
            result.append(group[0])
    return result
