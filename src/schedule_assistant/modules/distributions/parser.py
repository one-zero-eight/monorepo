"""Resilient Excel distribution table parser."""

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

EMAIL_HEADER_KEYWORDS = (
    "e-mail",
    "email",
    "mail",
    "доменн",
    "идентификатор",
)

MEMBERSHIP_HEADER_KEYWORDS = (
    "электив",
    "elective",
    "e group",
    "учебная группа",
    "group",
    "групп",
)

# Prefer elective columns over academic Group when both match.
MEMBERSHIP_PRIORITY_KEYWORDS = (
    "электив",
    "elective",
    "e group",
    "учебная группа",
)


@dataclass
class ParsedDistribution:
    sheet_names: list[str]
    sheet_name: str
    columns: list[str]
    header_row_index: int
    email_column: str | None
    membership_columns: list[str]
    forward_fill_columns: list[str]
    labels: list[DistributionLabel]
    sample_rows: list[dict[str, str]]
    row_count: int
    email_count: int
    emails_by_label: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class DistributionLabel:
    label: str
    email_count: int


def normalize_email(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    if not text:
        return None
    text = text.casefold()
    if not EMAIL_RE.match(text):
        return None
    return text


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _normalize_header(value: object, index: int) -> str:
    text = normalize_cell(value)
    if text:
        return text
    return f"Column {index + 1}"


def _header_score(cells: list[object]) -> int:
    non_empty = sum(1 for cell in cells if normalize_cell(cell))
    if non_empty < 2:
        return 0
    joined = " ".join(normalize_cell(cell).casefold() for cell in cells)
    score = non_empty
    if any(keyword in joined for keyword in EMAIL_HEADER_KEYWORDS):
        score += 10
    if any(keyword in joined for keyword in MEMBERSHIP_HEADER_KEYWORDS):
        score += 5
    return score


def _detect_header_row(rows: list[tuple[object, ...]]) -> int:
    best_index = 0
    best_score = -1
    limit = min(len(rows), 30)
    for index in range(limit):
        score = _header_score(list(rows[index]))
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _unique_column_names(raw_headers: list[object]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, raw in enumerate(raw_headers):
        base = _normalize_header(raw, index)
        count = seen.get(base, 0)
        seen[base] = count + 1
        names.append(base if count == 0 else f"{base} ({count + 1})")
    return names


def _column_looks_like_email(header: str) -> bool:
    lowered = header.casefold()
    return any(keyword in lowered for keyword in EMAIL_HEADER_KEYWORDS)


def _membership_priority(header: str) -> int:
    lowered = header.casefold()
    for index, keyword in enumerate(MEMBERSHIP_PRIORITY_KEYWORDS):
        if keyword in lowered:
            return index
    if any(keyword in lowered for keyword in MEMBERSHIP_HEADER_KEYWORDS):
        return len(MEMBERSHIP_PRIORITY_KEYWORDS)
    return 999


def _column_looks_like_membership(header: str) -> bool:
    return _membership_priority(header) < 999


def _email_hit_ratio(values: list[str]) -> float:
    non_empty = [value for value in values if value]
    if not non_empty:
        return 0.0
    hits = sum(1 for value in non_empty if normalize_email(value) is not None)
    return hits / len(non_empty)


def _is_sparse(values: list[str]) -> bool:
    non_empty = [value for value in values if value]
    if not non_empty:
        return False
    return len(non_empty) < len(values) * 0.7


def _should_forward_fill(header: str, values: list[str]) -> bool:
    """Forward-fill only sparse section-header columns (academic Group), not electives."""
    lowered = header.casefold()
    if any(keyword in lowered for keyword in ("электив", "elective", "e group", "email", "e-mail", "mail")):
        return False
    if "учебная" in lowered:
        return False
    if not _is_sparse(values):
        return False
    # Sparse generic Group / untitled first column used as section headers.
    return "group" in lowered or "групп" in lowered or lowered.startswith("column ") or not normalize_cell(header)


def _forward_fill(values: list[str]) -> list[str]:
    filled: list[str] = []
    last = ""
    for value in values:
        if value:
            last = value
        filled.append(last)
    return filled


def _fallback_membership_columns(
    columns: list[str],
    detected_email: str | None,
    column_values: dict[str, list[str]],
) -> list[str]:
    """When headers lack group keywords, pick categorical non-email columns."""
    candidates: list[str] = []
    for name in columns:
        if name == detected_email:
            continue
        if _email_hit_ratio(column_values[name]) >= 0.5:
            continue
        values = [value for value in column_values[name] if value]
        if not values:
            continue
        unique_count = len(set(values))
        # Prefer repeated labels over per-row unique names.
        if unique_count < len(values) and unique_count <= max(2, len(values) // 2):
            candidates.append(name)
    return candidates


def parse_distribution_xlsx(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
    email_column: str | None = None,
    membership_columns: list[str] | None = None,
    forward_fill_columns: list[str] | None = None,
) -> ParsedDistribution:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise ValueError("Workbook has no sheets")

        selected_sheet = sheet_name or sheet_names[0]
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {selected_sheet!r}")

        worksheet = workbook[selected_sheet]
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        raise ValueError("Sheet is empty")

    header_row_index = _detect_header_row(rows)
    header_row = list(rows[header_row_index])
    # Trim trailing empty header cells but keep at least one.
    while len(header_row) > 1 and not normalize_cell(header_row[-1]):
        header_row.pop()
    columns = _unique_column_names(header_row)
    width = len(columns)

    data_rows = rows[header_row_index + 1 :]
    column_values: dict[str, list[str]] = {name: [] for name in columns}
    for row in data_rows:
        padded = list(row) + [None] * max(0, width - len(row))
        for index, name in enumerate(columns):
            column_values[name].append(normalize_cell(padded[index] if index < len(padded) else None))

    detected_email = email_column
    if detected_email is None:
        header_candidates = [name for name in columns if _column_looks_like_email(name)]
        if header_candidates:
            detected_email = header_candidates[0]
        else:
            best_name = None
            best_ratio = 0.0
            for name in columns:
                ratio = _email_hit_ratio(column_values[name])
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_name = name
            if best_name is not None and best_ratio >= 0.5:
                detected_email = best_name

    detected_membership = membership_columns
    if detected_membership is None:
        candidates = [name for name in columns if name != detected_email and _column_looks_like_membership(name)]
        candidates.sort(key=_membership_priority)
        # If elective-like columns exist, drop generic "Group" academic columns from auto-detect.
        has_priority = any(_membership_priority(name) < len(MEMBERSHIP_PRIORITY_KEYWORDS) - 1 for name in candidates)
        if has_priority:
            candidates = [
                name
                for name in candidates
                if _membership_priority(name) < len(MEMBERSHIP_PRIORITY_KEYWORDS) - 1
                or "учебная" in name.casefold()
                or "e group" in name.casefold()
            ]
        if not candidates:
            candidates = _fallback_membership_columns(columns, detected_email, column_values)
        detected_membership = candidates
    else:
        detected_membership = [name for name in detected_membership if name in columns]

    detected_forward_fill = forward_fill_columns
    if detected_forward_fill is None:
        detected_forward_fill = [
            name for name in columns if name != detected_email and _should_forward_fill(name, column_values[name])
        ]
    else:
        detected_forward_fill = [name for name in detected_forward_fill if name in columns]

    working_values = {name: list(values) for name, values in column_values.items()}
    for name in detected_forward_fill:
        working_values[name] = _forward_fill(working_values[name])

    emails_by_label: dict[str, list[str]] = defaultdict(list)
    seen_by_label: dict[str, set[str]] = defaultdict(set)
    email_count = 0
    sample_rows: list[dict[str, str]] = []

    for row_index in range(len(data_rows)):
        row_dict = {name: working_values[name][row_index] for name in columns}
        email = normalize_email(row_dict.get(detected_email, "")) if detected_email else None
        if email is None:
            continue
        email_count += 1
        for membership_col in detected_membership:
            label = working_values[membership_col][row_index].strip()
            if not label:
                continue
            if email not in seen_by_label[label]:
                seen_by_label[label].add(email)
                emails_by_label[label].append(email)
        if len(sample_rows) < 8:
            sample = {"email": email}
            for membership_col in detected_membership:
                sample[membership_col] = working_values[membership_col][row_index]
            sample_rows.append(sample)

    labels = [
        DistributionLabel(label=label, email_count=len(emails))
        for label, emails in sorted(emails_by_label.items(), key=lambda item: (-len(item[1]), item[0].casefold()))
    ]

    return ParsedDistribution(
        sheet_names=sheet_names,
        sheet_name=selected_sheet,
        columns=columns,
        header_row_index=header_row_index,
        email_column=detected_email,
        membership_columns=detected_membership,
        forward_fill_columns=detected_forward_fill,
        labels=labels,
        sample_rows=sample_rows,
        row_count=len(data_rows),
        email_count=email_count,
        emails_by_label=dict(emails_by_label),
    )
