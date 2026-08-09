import datetime as dtm
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from src.schedule_assistant.modules.schedule.export_xlsx import export_schedule_xlsx
from src.schedule_assistant.modules.schedule_config.repository import ScheduleConfigRepository
from src.schedule_assistant.modules.schedule_config.schemas import (
    ComponentSessionSeries,
    CourseConfig,
    CoursesConfig,
    InstructorConfig,
    ScheduleConfig,
    SectionConfig,
    SessionOccurrence,
    StudentsGroups,
    TermConfig,
    WeeklyPatternSlot,
)
from src.schedule_assistant.weekday import Weekday


def _sample_config() -> ScheduleConfig:
    return ScheduleConfig(
        term=TermConfig(
            name="Fall 2026",
            semester=TermConfig.DateRange(
                start_date=dtm.date(2026, 9, 1),
                end_date=dtm.date(2026, 9, 30),
            ),
            sections=[
                SectionConfig(
                    code="core",
                    name="Core",
                    kind="core",
                    programs=[
                        SectionConfig.SectionProgram(
                            code="BS_Y1",
                            name="BS Y1",
                            tracks=[
                                SectionConfig.SectionProgram.ProgramTrack(
                                    code="CSE",
                                    name="Software Engineering",
                                    groups=["B25-CSE-01", "B25-CSE-02"],
                                )
                            ],
                        )
                    ],
                ),
                SectionConfig(
                    code="electives",
                    name="Electives",
                    kind="electives",
                    default_layout="calendar",
                    programs=[
                        SectionConfig.SectionProgram(
                            code="TECH",
                            name="Tech",
                            groups=["ELEC-01"],
                        )
                    ],
                ),
            ],
        ),
        instructors=[
            InstructorConfig.Instructor(
                id="t1",
                name_en="Alice Teacher",
                name_ru="Алиса",
                email="t1@x.ru",
                alias="alice",
                position="Professor",
            ),
            InstructorConfig.Instructor(
                id="t2",
                name_en="Bob Assistant",
                email="t2@x.ru",
                position="Teaching Assistant",
            ),
        ],
        students_groups=[
            StudentsGroups(
                code="B25-CSE-01",
                kind="core",
                name="B25-CSE-01",
                estimated_size=27,
                students=["b.student@innopolis.university", "a.student@innopolis.university"],
            ),
            StudentsGroups(
                code="B25-CSE-02",
                kind="core",
                name="B25-CSE-02",
                estimated_size=26,
                students=["c.student@innopolis.university"],
            ),
            StudentsGroups(
                code="ELEC-01",
                kind="elective",
                name="SRE Course",
                students=["e.student@innopolis.university", "d.student@innopolis.university"],
            ),
        ],
        courses=[
            CourseConfig(
                name="Math I",
                short_name="MATH",
                name_ru="Математика I",
                instructors=[
                    CourseConfig.CourseInstructor(id="t1", role="Primary"),
                    CourseConfig.CourseInstructor(id="t2", role="Teaching Assistant"),
                ],
                components=[
                    CourseConfig.Component(
                        tag="lec",
                        student_groups=["@BS_Y1/Software Engineering"],
                        sessions=[
                            ComponentSessionSeries(
                                audience=["@BS_Y1/Software Engineering"],
                                weekly_pattern=[
                                    WeeklyPatternSlot(
                                        weekday=Weekday.MONDAY,
                                        start_time=dtm.time(9, 0),
                                        end_time=dtm.time(10, 30),
                                        room="108",
                                        instructor="t1",
                                    )
                                ],
                            )
                        ],
                    )
                ],
            ),
            CourseConfig(
                name="SRE Course",
                short_name="SRE",
                instructors=[
                    CourseConfig.CourseInstructor(id="t1", role="Primary"),
                ],
                components=[
                    CourseConfig.Component(
                        tag="lab",
                        student_groups=["ELEC-01"],
                        sessions=[
                            ComponentSessionSeries(
                                audience=["ELEC-01"],
                                occurrences=[
                                    SessionOccurrence(
                                        date=dtm.date(2026, 9, 4),
                                        start_time=dtm.time(12, 40),
                                        end_time=dtm.time(14, 10),
                                        room="209",
                                        instructor="t1",
                                    )
                                ],
                            )
                        ],
                    )
                ],
            ),
        ],
    )


def test_export_has_sheet_per_section_with_default_layouts() -> None:
    data, filename = export_schedule_xlsx(_sample_config())
    assert filename == "Fall 2026.xlsx"
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Core", "Electives", "Distributions", "Instructors", "Subjects"]

    core = wb["Core"]
    assert core["B1"].value == "BS Y1"
    assert core["B2"].value == "B25-CSE-01 (27)"
    assert core["C2"].value == "B25-CSE-02 (26)"
    assert core["A3"].value == "MONDAY"
    assert core["A4"].value == "09:00-10:30"
    assert core["B4"].value == "Math I (lec)"
    assert core["B5"].value == "Alice Teacher"
    assert core["B6"].value == "108"
    merges = {str(item) for item in core.merged_cells.ranges}
    assert "B4:C4" in merges
    assert "B5:C5" in merges
    assert "B6:C6" in merges
    assert "A4:A6" in merges
    assert "A1:A2" in merges
    monday_slots = [
        core.cell(r, 1).value for r in range(4, 4 + 3 * len(_sample_config().term.time_slots)) if core.cell(r, 1).value
    ]
    assert monday_slots == [
        f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')}" for s in _sample_config().term.time_slots
    ]

    electives = wb["Electives"]
    assert electives["B1"].value == "Monday"
    assert electives["A2"].value == "Week 1"
    found = False
    for row in electives.iter_rows(min_row=1, max_row=40, max_col=8):
        for cell in row:
            if cell.value and "SRE" in str(cell.value) and "209" in str(cell.value):
                found = True
                break
    assert found

    instructors = wb["Instructors"]
    assert [instructors.cell(1, col).value for col in range(1, 8)] == [
        "Name EN",
        "Name RU",
        "Email",
        "Alias",
        "Position",
        "Courses",
        "ID",
    ]
    assert instructors["A2"].value == "Alice Teacher"
    assert instructors["B2"].value == "Алиса"
    assert instructors["C2"].value == "t1@x.ru"
    assert instructors["D2"].value == "alice"
    assert instructors["E2"].value == "Professor"
    assert instructors["F2"].value == "Math I (Primary)\nSRE Course (Primary)"
    assert instructors["G2"].value == "t1"
    assert instructors["A3"].value == "Bob Assistant"
    assert instructors["F3"].value == "Math I (Teaching Assistant)"
    assert instructors["G3"].value == "t2"

    subjects = wb["Subjects"]
    assert [subjects.cell(1, col).value for col in range(1, 7)] == [
        "Section",
        "Short name",
        "Name",
        "Name RU",
        "Groups",
        "Instructors",
    ]
    assert subjects["A2"].value == "Core"
    assert subjects["B2"].value == "MATH"
    assert subjects["C2"].value == "Math I"
    assert subjects["D2"].value == "Математика I"
    assert subjects["E2"].value == "BS_Y1/Software Engineering: B25-CSE-01, B25-CSE-02"
    assert subjects["F2"].value == "Alice Teacher (Primary)\nBob Assistant (Teaching Assistant)"
    assert subjects["A3"].value == "Electives"
    assert subjects["B3"].value == "SRE"
    assert subjects["C3"].value == "SRE Course"
    assert subjects["E3"].value == "ELEC-01"
    assert subjects["F3"].value == "Alice Teacher (Primary)"

    distributions = wb["Distributions"]
    assert distributions["A1"].value == "E-mail"
    assert distributions["B1"].value == "Group"
    assert distributions["C1"].value == "Section"
    assert distributions["A2"].value == "a.student@innopolis.university"
    assert distributions["B2"].value == "B25-CSE-01"
    assert distributions["C2"].value == "Core"
    assert distributions["A3"].value == "b.student@innopolis.university"
    assert distributions["A4"].value == "c.student@innopolis.university"
    assert distributions["B4"].value == "B25-CSE-02"
    assert distributions["A5"].value == "d.student@innopolis.university"
    assert distributions["B5"].value == "SRE Course"
    assert distributions["C5"].value == "Electives"
    assert distributions["A6"].value == "e.student@innopolis.university"


def test_subjects_groups_track_only_when_complete() -> None:
    config = _sample_config()
    config.courses[0].components[0].student_groups = ["B25-CSE-01"]
    session = config.courses[0].components[0].sessions
    assert session is not None
    session[0].audience = ["B25-CSE-01"]
    data, _ = export_schedule_xlsx(config)
    subjects = load_workbook(io.BytesIO(data))["Subjects"]
    assert subjects["C2"].value == "Math I"
    assert subjects["E2"].value == "B25-CSE-01"


def test_subjects_section_block_has_outer_border() -> None:
    config = _sample_config()
    config.courses.append(
        CourseConfig(
            name="Physics I",
            short_name="PHYS",
            instructors=[CourseConfig.CourseInstructor(id="t1", role="Primary")],
            components=[
                CourseConfig.Component(
                    tag="lec",
                    student_groups=["@BS_Y1/Software Engineering"],
                    sessions=[
                        ComponentSessionSeries(
                            audience=["@BS_Y1/Software Engineering"],
                            weekly_pattern=[
                                WeeklyPatternSlot(
                                    weekday=Weekday.TUESDAY,
                                    start_time=dtm.time(9, 0),
                                    end_time=dtm.time(10, 30),
                                    room="108",
                                    instructor="t1",
                                )
                            ],
                        )
                    ],
                )
            ],
        )
    )
    data, _ = export_schedule_xlsx(config)
    subjects = load_workbook(io.BytesIO(data))["Subjects"]
    assert subjects["A2"].value == "Core"
    assert subjects["A3"].value == "Core"
    assert subjects["C2"].value == "Math I"
    assert subjects["C3"].value == "Physics I"
    # Outer box around Core rows 2–3 is thick; internal edges are thin.
    assert subjects["A2"].border.top is not None and subjects["A2"].border.top.style == "thick"
    assert subjects["A2"].border.bottom is not None and subjects["A2"].border.bottom.style == "thin"
    assert subjects["A3"].border.top is not None and subjects["A3"].border.top.style == "thin"
    assert subjects["A3"].border.bottom is not None and subjects["A3"].border.bottom.style == "thick"
    assert subjects["A2"].border.left is not None and subjects["A2"].border.left.style == "thick"
    assert subjects["B2"].border.left is not None and subjects["B2"].border.left.style == "thin"
    assert subjects["F2"].border.right is not None and subjects["F2"].border.right.style == "thick"
    assert subjects["A4"].value == "Electives"
    assert subjects["A4"].border.top is not None and subjects["A4"].border.top.style == "thick"
    assert subjects["A4"].border.bottom is not None and subjects["A4"].border.bottom.style == "thick"
    assert subjects.sheet_view.showGridLines is False


def test_groups_export_grows_row_height_for_long_wrapped_text() -> None:
    config = _sample_config()
    config.term.sections = [config.term.sections[0]]
    for i in range(8):
        config.instructors.append(
            InstructorConfig.Instructor(
                id=f"t{i + 2}",
                name_en=f"Instructor Number {i + 2} Longname",
                email=f"t{i + 2}@x.ru",
            )
        )
    session = config.courses[0].components[0].sessions
    assert session is not None
    pattern = session[0].weekly_pattern
    assert pattern is not None
    pattern[0].instructor = [f"t{i + 1}" for i in range(9)]
    data, _ = export_schedule_xlsx(config)
    ws = load_workbook(io.BytesIO(data)).active
    assert ws is not None
    assert (ws.row_dimensions[4].height or 0) >= 15
    assert (ws.row_dimensions[5].height or 0) > 15
    assert (ws.row_dimensions[5].height or 0) >= 30


def test_groups_export_repeats_time_column_per_program() -> None:
    config = _sample_config()
    config.term.sections = [config.term.sections[0]]
    config.term.sections[0].programs.append(
        SectionConfig.SectionProgram(
            code="BS_Y2",
            name="BS Y2",
            groups=["B24-CSE-01"],
        )
    )
    config.students_groups.append(
        StudentsGroups(code="B24-CSE-01", kind="core", name="B24-CSE-01", estimated_size=29),
    )
    config.courses.append(
        CourseConfig(
            name="Math II",
            short_name="MATH2",
            components=[
                CourseConfig.Component(
                    tag="lec",
                    student_groups=["B24-CSE-01"],
                    sessions=[
                        ComponentSessionSeries(
                            audience=["B24-CSE-01"],
                            weekly_pattern=[
                                WeeklyPatternSlot(
                                    weekday=Weekday.MONDAY,
                                    start_time=dtm.time(9, 0),
                                    end_time=dtm.time(10, 30),
                                    room="201",
                                    instructor="t1",
                                )
                            ],
                        )
                    ],
                )
            ],
        )
    )
    data, _ = export_schedule_xlsx(config)
    ws = load_workbook(io.BytesIO(data)).active
    assert ws is not None
    assert ws["B1"].value == "BS Y1"
    assert ws["E1"].value == "BS Y2"
    assert ws["A3"].value == "MONDAY"
    assert ws["D3"].value == "MONDAY"
    assert ws["A4"].value == "09:00-10:30"
    assert ws["D4"].value == "09:00-10:30"
    assert ws["E4"].value == "Math II (lec)"
    assert ws.max_column == 5
    merges = {str(item) for item in ws.merged_cells.ranges}
    assert "A4:A6" in merges
    assert "D4:D6" in merges
    assert "D1:D2" in merges


def _seed_export_config(repo: ScheduleConfigRepository) -> None:
    config = _sample_config()
    term_without_sections = config.term.model_copy(update={"sections": []})
    repo.set_term(term_without_sections, saved_by="mod@innopolis.university")
    repo.set_instructors(InstructorConfig(instructors=config.instructors), saved_by="mod@innopolis.university")
    from src.schedule_assistant.modules.schedule_config.schemas import RoomConfig, SectionsConfig

    repo.create_room(RoomConfig.Room(id="108", name="108", capacity=100), saved_by="mod@innopolis.university")
    repo.create_room(RoomConfig.Room(id="209", name="209", capacity=40), saved_by="mod@innopolis.university")
    repo.set_sections(
        SectionsConfig(sections=config.term.sections, students_groups=config.students_groups),
        saved_by="mod@innopolis.university",
    )
    repo.set_courses(CoursesConfig(courses=config.courses), saved_by="mod@innopolis.university")


@pytest.mark.asyncio
async def test_export_xlsx_api(
    authenticated_client: AsyncClient,
    schedule_data_repo: ScheduleConfigRepository,
) -> None:
    _seed_export_config(schedule_data_repo)

    response = await authenticated_client.get("/schedule/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    disposition = response.headers["content-disposition"]
    assert "Fall 2026.xlsx" in disposition or "Fall%202026.xlsx" in disposition
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Core", "Electives", "Distributions", "Instructors", "Subjects"]
    assert wb["Core"]["B4"].value == "Math I (lec)"
    assert wb["Electives"]["B1"].value == "Monday"
    assert wb["Instructors"]["G2"].value == "t1"
    assert wb["Subjects"]["A2"].value == "Core"
    assert wb["Subjects"]["C2"].value == "Math I"
    assert wb["Subjects"]["E2"].value == "BS_Y1/Software Engineering: B25-CSE-01, B25-CSE-02"
    assert wb["Subjects"]["F2"].value == "Alice Teacher (Primary)\nBob Assistant (Teaching Assistant)"
    assert wb["Distributions"]["A1"].value == "E-mail"
    assert wb["Distributions"]["A2"].value == "a.student@innopolis.university"
